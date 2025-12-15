## forecast_system.py

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

# -----------------------------
# Forecast Function
# -----------------------------
def run_forecast(input_file):
    """
    Reads historical data, runs a simple forecast (placeholder), and returns forecast DataFrame.
    Replace this with your actual forecasting logic.
    """
    # Load historical data
    df = pd.read_csv(input_file)

    # Example: Simple rolling average forecast for next 13 months
    # Replace this with your real forecasting model
    last_value = df['Volume'].iloc[-1]
    forecast_values = [last_value * (1 + 0.02 * i) for i in range(1, 14)]
    months = pd.date_range(start=pd.Timestamp.today(), periods=13, freq='M').strftime("%b-%Y")

    forecast_df = pd.DataFrame({
        'Month': months,
        'Forecast Volume': forecast_values
    })

    return forecast_df

# -----------------------------
# Excel Writing Function
# -----------------------------
def write_forecast_to_excel(forecast_df, output_path):
    """
    Writes the forecast DataFrame to a nicely formatted Excel file.
    """
    # Ensure the output folder exists
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "13 Month Forecast"

    # Add title
    ws["A1"] = "13-Month Volume Forecast"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])  # Blank row

    # Add DataFrame to sheet
    for r_idx, row in enumerate(dataframe_to_rows(forecast_df.round(2), index=False, header=True), start=1):
        ws.append(row)
        # Bold header
        if r_idx == 1:
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 4

    # Save Excel file
    wb.save(output_path)
    print(f"Forecast written to {output_path}")

# -----------------------------
# Main Script
# -----------------------------
if __name__ == "__main__":
    # Input CSV file path (replace with your actual path)
    input_csv = "historical_data.csv"

    # Run forecast
    forecast_df = run_forecast(input_csv)

    # Output Excel path
    output_file = "forecast_outputs/13_month_volume_forecast.xlsx"

    # Write to Excel
    write_forecast_to_excel(forecast_df, output_file)
